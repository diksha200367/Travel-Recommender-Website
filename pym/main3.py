import pymongo
from openpyxl import load_workbook
from collections import defaultdict

# MongoDB connection
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client['Cities3']
collection = db['mySampleCollectionForCities3']

# Path to the Excel file
excel_file = "D:\\Data_Desc_Season.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active

# Mapping of cities to their corresponding custom IDs
city_id_mapping = {
    "AGRA": 1,
    "AHMEDABAD": 2,
    "AJMER": 3,
    "ALAPPUZHA": 4,
    "AMRITSAR": 5,
    "BANARAS": 6,
    "BANGLORE": 7,
    "BHOPAL": 8,
    "BHUBANESHWAR": 9,
    "CHANDIGARH": 10,
    "CHENNAI": 11,
    "CHITTORGARH": 12,
    "COIMBATORE": 13,
    "COORG": 14,
    "DARJEELING": 15,
    "DEHRADHUN": 16,
    "DELHI": 17,
    "DEOGHAR": 18,
    "DHARAMSHALA": 19,
    "GANGTOK": 20,
    "GOA": 21,
    "GULMARG": 22,
    "GUWAHATI": 23,
    "GWALIOR": 24,
    "HARIDWAR": 25,
    "HUMPY": 26,
    "HYDERABAD": 27,
    "INDORE": 28,
    "JAIPUR": 29,
    "JAISALMER": 30,
    "JODHPUR": 31,
    "KANCHIPURAM": 32,
    "KANNUR": 33,
    "KANPUR": 34,
    "KHAJURAHO": 35,
    "KOCHI": 36,
    "KODAIKANAL": 37,
    "KOLKATA": 38,
    "LEH": 39,
    "LUCKNOW": 40,
    "MADHURAI": 41,
    "MAHABALIPURAM": 42,
    "MANALI": 43,
    "MATHURA": 44,
    "MUMBAI": 45,
    "MUNNAR": 46,
    "MUSSOORIE": 47,
    "MYSORE": 48,
    "NAINITAAL": 49,
    "NASHIK": 50,
    "OOTI": 51,
    "PAHALGAM": 52,
    "PONDICHERRY": 53,
    "PUNE": 54,
    "PURI": 55,
    "RISHIKESH": 56,
    "RANN OF KUTCH": 57,
    "RANTHAMBORE": 58,
    "RAMESHWARAM": 59,
    "SHILLONG": 60,
    "SHIMLA": 61,
    "SRINAGAR": 62,
    "UDAIPUR": 63,
    "UJJAIN": 64,
    "VRINDAVAN": 65,
    "WAYNAD": 66,
    "KANYAKUMARI": 67
}

# Group data by city
city_data = defaultdict(lambda: {"STATE": "", "DESTINATIONS": []})

# Loop through the rows in the Excel sheet and prepare the data
for row in sheet.iter_rows(min_row=2, values_only=True):
    city, state, tourist_spot, attraction_type, season, activities = row

    # Check if the city is in the mapping, and assign the custom _id if it is
    if city:
        city_id = city_id_mapping.get(city.upper(), None)

        # Assign the state if it hasn't been set yet
        if not city_data[city]["STATE"]:
            city_data[city]["STATE"] = state

        # Append destination details to the city's destinations list
        city_data[city]["DESTINATIONS"].append({
            "TOURIST SPOT": tourist_spot,
            "TYPE OF ATTRACTION": attraction_type,
            "SEASON": season,
            "ACTIVITIES": [activity.strip() for activity in activities.split(",") if activity] if activities else []
        })

# Convert the grouped data to a format suitable for MongoDB insertion
formatted_data = [
    {
        "_id": city_id_mapping[city.upper()],
        "CITY": city,
        "STATE": details["STATE"],
        "DESTINATIONS": details["DESTINATIONS"],
    }
    for city, details in city_data.items()
    if city.upper() in city_id_mapping
]

# Clear existing data and insert formatted data into MongoDB
collection.delete_many({})

if formatted_data:
    try:
        result = collection.insert_many(formatted_data, ordered=False)  # `ordered=False` allows partial inserts in case of errors
        print(f"Inserted {len(result.inserted_ids)} documents into MongoDB.")
    except Exception as e:
        print(f"Error inserting documents: {e}")
else:
    print("No data to insert.")
